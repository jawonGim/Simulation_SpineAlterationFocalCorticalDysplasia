from matplotlib import pyplot as plt

from openpyxl import Workbook, load_workbook
import os

def EPSP_multipoints(all=True, model="ctl", alter="head", N=100, show=True):
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    fig, ax = plt.subplots()
    ax.set_aspect(1.)
    ax = fig.add_axes([0, 0, 1, 1])  # x0, y0, width, height (fill 100%)

    accum_vhead = h.Vector()
    accum_vbase = h.Vector()
    accum_vsoma = h.Vector()
    for i in range(1, N+1) :
        cellA = cell.cell(name="control")
        if model == "ctl" :
            cellA.create_cell()
        else :
            if alter == "head" :
                cellspec = dict()
                cellspec["spine_head_diam"] = 0.718
                cellspec["spine_len"] = 1.823-0.551+0.718
                cellA.create_cell_with_spec(cellspec)
            elif alter == "neck" :
                cellspec = dict()
                cellspec["spine_neck_diam"] = 0.301
                cellA.create_cell_with_spec(cellspec)

        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        t = h.Vector().record(h._ref_t)
        vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
        vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

        cellA.set_cell_properties(sf=2.0)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        if all : 
            plt.plot(t, vhead, linewidth=1, color="mistyrose")
            plt.plot(t, vbase, linewidth=1, color="lavender")
        
        plt.plot(t, cellA.v_soma, linewidth=1, color="lightgrey")

        if accum_vhead.size() == 0 :
            accum_vhead = vhead
            accum_vbase = vbase
            accum_vsoma = cellA.v_soma
        else :
            accum_vhead = accum_vhead.add(vhead)    
            accum_vbase = accum_vbase.add(vbase) 
            accum_vsoma = accum_vsoma.add(cellA.v_soma)

        del cellA

    if all :
        plt.plot(t, accum_vhead.div(N), linewidth=2.5, color="red")
        plt.plot(t, accum_vbase.div(N), linewidth=2.5, color="blue")
    plt.plot(t, accum_vsoma.div(N), linewidth=2.5, color="black")

    plt.xlim((265, 500))

    if all :
        min_y = -70.6
        max_y = min_y + 10
    else :
        min_y = -70.27
        max_y = min_y + 0.5 #soma only

    plt.ylim([min_y, max_y])  

    plt.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=False,      # ticks along the bottom edge are off
        top=False, 
        labelbottom=False)
    plt.tick_params(
        axis='y',          # changes apply to the y-axis
        which='both',      # both major and minor ticks are affected
        left=False,      # ticks along the bottom edge are off
        right=False,
        labelleft=False)

    #add scale bar
    from mpl_toolkits.axes_grid1.anchored_artists import AnchoredSizeBar
    xtick_ms = 20
    asb = AnchoredSizeBar(ax.transData,
                          size=xtick_ms,
                          label="",
                          loc='lower left',
                          borderpad=1, sep=3, size_vertical=0.002, 
                          frameon=False)
    ax.add_artist(asb)

    if all :
        ytick_mV = 1
    else :
        ytick_mV = 0.05 
    asb = AnchoredSizeBar(ax.transData,
                          size=0.1,
                          label="",
                          loc='lower left',
                          borderpad=1, sep=3, size_vertical=ytick_mV,
                          frameon=False)
    ax.add_artist(asb)
    if all :
        save_name = f"{alter}_{model}_EPSP_{xtick_ms}ms_{ytick_mV}mV_all.png"
    else :
        save_name = f"{alter}_{model}_EPSP_{xtick_ms}ms_{ytick_mV}mV_soma.png"

    # Force all spines OFF
    for spine in ax.spines.values():
        spine.set_visible(False)
        spine.set_color('white')

    plt.savefig(save_name, dpi=600, bbox_inches='tight', pad_inches=0, transparent=False)

    if show:
        plt.show()

def EPSP_multipoints_density(all=True, model="ctl", N=100, show=True):
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    fig, ax = plt.subplots()
    ax.set_aspect(1.)
    ax = fig.add_axes([0, 0, 1, 1])  # x0, y0, width, height (fill 100%)

    accum_vhead = h.Vector()
    accum_vbase = h.Vector()
    accum_vsoma = h.Vector()

    for i in range(1, N+1) :
        cellA = cell.cell(name="control")
        cellA.create_cell()

        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        t = h.Vector().record(h._ref_t)
        vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
        vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

        if model == "ctl" :
            cellA.set_cell_properties(sf=2.0)
        else :
            cellA.set_cell_properties(sf=1.2)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        if all : 
            plt.plot(t, vhead, linewidth=1, color="mistyrose")
            plt.plot(t, vbase, linewidth=1, color="lavender")
        
        plt.plot(t, cellA.v_soma, linewidth=1, color="lightgrey")

        if accum_vhead.size() == 0 :
            accum_vhead = vhead
            accum_vbase = vbase
            accum_vsoma = cellA.v_soma
        else :
            accum_vhead = accum_vhead.add(vhead)    
            accum_vbase = accum_vbase.add(vbase) 
            accum_vsoma = accum_vsoma.add(cellA.v_soma)

        del cellA

    if all :
        plt.plot(t, accum_vhead.div(N), linewidth=2.5, color="red")
        plt.plot(t, accum_vbase.div(N), linewidth=2.5, color="blue")
    plt.plot(t, accum_vsoma.div(N), linewidth=2.5, color="black")

    plt.xlim((265, 500))

    if all :
        min_y = -70.6
        if model == "fcd" :
            min_y = -70.7

        max_y = min_y + 10
    else :
        min_y = -70.274
        if model == "fcd" :
            min_y = -70.39

        max_y = min_y + 0.5 #soma only

    plt.ylim([min_y, max_y])  

    plt.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=False,      # ticks along the bottom edge are off
        top=False, 
        labelbottom=False)
    plt.tick_params(
        axis='y',          # changes apply to the y-axis
        which='both',      # both major and minor ticks are affected
        left=False,      # ticks along the bottom edge are off
        right=False,
        labelleft=False)

    #add scale bar
    from mpl_toolkits.axes_grid1.anchored_artists import AnchoredSizeBar
    xtick_ms = 20
    asb = AnchoredSizeBar(ax.transData,
                          size=xtick_ms,
                          label="",
                          loc='lower left',
                          borderpad=1, sep=3, size_vertical=0.002, 
                          frameon=False)
    ax.add_artist(asb)

    if all :
        ytick_mV = 1
    else :
        ytick_mV = 0.05 
    asb = AnchoredSizeBar(ax.transData,
                          size=0.1,
                          label="",
                          loc='lower left',
                          borderpad=1, sep=3, size_vertical=ytick_mV,
                          frameon=False)
    ax.add_artist(asb)
    if all :
        save_name = f"density_{model}_EPSP_{xtick_ms}ms_{ytick_mV}mV_all.png"
    else :
        save_name = f"density_{model}_EPSP_{xtick_ms}ms_{ytick_mV}mV_soma.png"
    
    # Force all spines on
    for spine in ax.spines.values():
        spine.set_visible(False)
        spine.set_color('white')

    plt.savefig(save_name, dpi=600, bbox_inches='tight', pad_inches=0, transparent=False)

    if show:
        plt.show()


def EPSP_multipoints_allfactors(all=True, model="ctl", N=100, show=True):
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    fig, ax = plt.subplots()
    ax.set_aspect(1.)
    ax = fig.add_axes([0, 0, 1, 1])  # x0, y0, width, height (fill 100%)

    accum_vhead = h.Vector()
    accum_vbase = h.Vector()
    accum_vsoma = h.Vector()

    for i in range(1, N+1) :
        cellA = cell.cell(name="control")
        if model == "ctl" :
            cellA.create_cell()
        else :
            cellspec = dict()
            cellspec["spine_head_diam"] = 0.718
            cellspec["spine_neck_diam"] = 0.301
            cellspec["spine_len"] = 1.696
            cellA.create_cell_with_spec(cellspec)


        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        t = h.Vector().record(h._ref_t)
        vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
        vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

        if model == "ctl" :
            cellA.set_cell_properties(sf=2.0)
        else :
            cellA.set_cell_properties(sf=1.2)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        if all : 
            plt.plot(t, vhead, linewidth=1, color="mistyrose")
            plt.plot(t, vbase, linewidth=1, color="lavender")
        
        plt.plot(t, cellA.v_soma, linewidth=1, color="lightgrey")

        if accum_vhead.size() == 0 :
            accum_vhead = vhead
            accum_vbase = vbase
            accum_vsoma = cellA.v_soma
        else :
            accum_vhead = accum_vhead.add(vhead)    
            accum_vbase = accum_vbase.add(vbase) 
            accum_vsoma = accum_vsoma.add(cellA.v_soma)

        del cellA

    if all :
        plt.plot(t, accum_vhead.div(N), linewidth=2.5, color="red")
        plt.plot(t, accum_vbase.div(N), linewidth=2.5, color="blue")
    plt.plot(t, accum_vsoma.div(N), linewidth=2.5, color="black")

    plt.xlim((265, 500))

    if all :
        min_y = -70.6
        if model == "fcd" :
            min_y = -70.7

        max_y = min_y + 10
    else :
        min_y = -70.274
        if model == "fcd" :
            min_y = -70.39

        max_y = min_y + 0.5 #soma only

    plt.ylim([min_y, max_y])  

    plt.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=False,      # ticks along the bottom edge are off
        top=False, 
        labelbottom=False)
    plt.tick_params(
        axis='y',          # changes apply to the y-axis
        which='both',      # both major and minor ticks are affected
        left=False,      # ticks along the bottom edge are off
        right=False,
        labelleft=False)

    #add scale bar
    from mpl_toolkits.axes_grid1.anchored_artists import AnchoredSizeBar
    xtick_ms = 20
    asb = AnchoredSizeBar(ax.transData,
                          size=xtick_ms,
                          label="",
                          loc='lower left',
                          borderpad=1, sep=3, size_vertical=0.002, 
                          frameon=False)
    ax.add_artist(asb)

    if all :
        ytick_mV = 1
    else :
        ytick_mV = 0.05 
    asb = AnchoredSizeBar(ax.transData,
                          size=0.1,
                          label="",
                          loc='lower left',
                          borderpad=1, sep=3, size_vertical=ytick_mV,
                          frameon=False)
    ax.add_artist(asb)
    if all :
        save_name = f"allfactor_{model}_EPSP_{xtick_ms}ms_{ytick_mV}mV_all.png"
    else :
        save_name = f"allfactor_{model}_EPSP_{xtick_ms}ms_{ytick_mV}mV_soma.png"

    # Force all spines on
    for spine in ax.spines.values():
        spine.set_visible(False)
        spine.set_color('white')

    plt.savefig(save_name, dpi=600, bbox_inches='tight', pad_inches=0, transparent=False)

    if show:
        plt.show()


## HERE to call the functions 


EPSP_multipoints_allfactors(all=True, model="fcd", show=False)
EPSP_multipoints_allfactors(all=False, model="fcd", show=False)

# #about density changes only
# EPSP_multipoints_density(model="ctl", show=False)   
# EPSP_multipoints_density(model="fcd", show=False)
# EPSP_multipoints_density(all=False, model="ctl", show=False)    #for somatic EPSP only
# EPSP_multipoints_density(all=False, model="fcd", show=False)    #for somatic EPSP only

# #about spine head diameter changes only
# EPSP_multipoints(model="ctl", alter="head", show=False)
# EPSP_multipoints(model="fcd", alter="head", show=False)
# EPSP_multipoints(all=False, model="ctl", alter="head", show=False)  #for somatic EPSP only
# EPSP_multipoints(all=False, model="fcd", alter="head", show=False)  #for somatic EPSP only

# #about spine neck diameter changes only
# EPSP_multipoints(model="ctl", alter="neck", show=False)
# EPSP_multipoints(model="fcd", alter="neck", show=False)
# EPSP_multipoints(all=False, model="ctl", alter="neck", show=False)  #for somatic EPSP only
# EPSP_multipoints(all=False, model="fcd", alter="neck", show=False)  #for somatic EPSP only