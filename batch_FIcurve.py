from neuron import gui, h
from neuron.units import ms, sec, mV
import morphology

import time
import random
from utils import measure_input_resistance
from openpyxl import Workbook, load_workbook
import os
import math

def measure_firing_rate(cell, stim_hz, vinit=-70*mV, dt=0.1*ms, tstop=12000*ms, weights=(31, 14)) :
    ampa_weight = weights[0]
    nmda_weight = weights[1]
    
    # stim_start = 0 *ms
    stim_intv = 1_000 * ms / stim_hz
    stim_num = tstop / stim_intv
    drop_both = stim_intv*2

    drop_both = min(drop_both, 500*ms)  # ensure drop_both maximum is 500ms

    random.seed(time.time())
    stims = []
    ncl_ampa = []
    ncl_nmda = []
    #set stimulation 
    for ampa, nmda in zip(cell.ampar, cell.nmdar) :
        stims.append(h.NetStim())
        stims[-1].start = random.randrange(0, int(stim_intv))
        stims[-1].interval = stim_intv
        stims[-1].number = stim_num 
        stims[-1].noise = 1 

        ncl_ampa.append(h.NetCon(stims[-1], ampa))    
        ncl_ampa[-1].weight[0] = ampa_weight
        ncl_ampa[-1].delay = 0 * ms
    
        ncl_nmda.append(h.NetCon(stims[-1], nmda))
        ncl_nmda[-1].weight[0] = nmda_weight
        ncl_nmda[-1].delay = 0 * ms

    h.dt = dt
    h.v_init = vinit
    h.finitialize(vinit)
    h.continuerun(tstop)

    firing = list(cell.spike_times)

    cropped = [sp for sp in firing if sp < tstop-drop_both and sp > drop_both]
    if len(cropped) == 0 :
        fr = 0
    else :
        fr = len(cropped)/(tstop-drop_both*2)*sec
    print(f"firing rate: {fr}")
    return fr



spec_ctl = dict()

spec_fcd = dict()
spec_fcd["nspines_perdend"] = 30
spec_fcd["spine_len"] = 1.696
spec_fcd["spine_head_diam"] = 0.718
spec_fcd["spine_neck_diam"] = 0.301 
spec_fcd["model"] = "epilepsy" 

vinit = -70 * mV

ntrial=1
ninput_hz = 25
ndend_comb = 10

xls_fname = "FI_for_various_dendcomb.xlsx"
if os.path.isfile(xls_fname) :
    wb = load_workbook(filename=xls_fname)
else :
    wb = Workbook()
ws = wb.active

if ws.max_row < 2 :
    ridx = 1
    ws.cell(row=ridx, column=1, value="napical")
    ws.cell(row=ridx, column=2, value="nbasal")
    ws.cell(row=ridx, column=3, value="Ri_CTL")
    ws.cell(row=ridx, column=4, value="Ri_FCD")
    ws.cell(row=ridx, column=5, value="I_hz")
    ws.cell(row=ridx, column=6, value="trial")
    ws.cell(row=ridx, column=7, value="fr_CTL")
    ws.cell(row=ridx, column=8, value="fr_FCD")
    nbasal_from = 0
    inputhz_from = 0
    trial_from = 1
else :
    ridx = ws.max_row -1
    #compute loop index from ridx
    trial_from = math.ceil(ridx%ntrial)+1
    nbasal_from = ridx//ntrial//ninput_hz
    inputhz_from = math.floor(ridx/ntrial - ninput_hz*nbasal_from)

    ridx = ws.max_row


for nbasal_ in range(nbasal_from, ndend_comb) :
    nbasal = 1 + nbasal_
    print(f"nbasal: {nbasal}")

    spec_ctl["nbasal"] = nbasal
    spec_fcd["nbasal"] = nbasal

    for input_hz_ in range(inputhz_from, ninput_hz) :
        input_hz = 0.2 + 0.4*input_hz_
        if input_hz > 5 :
            break
        print(f"input_hz: {input_hz}")
        
        for trial in range(trial_from, ntrial+1) :
            print(f"nbasal: {nbasal}, input_hz: {input_hz}, trial: {trial}")
            ridx += 1

            ws.cell(row=ridx, column=1, value=20-nbasal)
            ws.cell(row=ridx, column=2, value=nbasal)
            ws.cell(row=ridx, column=5, value=input_hz)

            ws.cell(row=ridx, column=6, value=trial)

            #control cell
            cell = morphology.cell(verbose=False)
            cell.create_cell_with_spec(cellspec=spec_ctl)
            cell.set_cell_properties()
        
            h.celsius = 37

            ri = measure_input_resistance(cell, vinit)
            fr = measure_firing_rate(cell, stim_hz=input_hz, vinit=vinit, dt=0.1, weights=(31,14))

            ws.cell(row=ridx, column=3, value=ri)
            ws.cell(row=ridx, column=7, value=fr)
            del cell

            #epilepsy cell
            cell = morphology.cell(verbose=False)
            cell.create_cell_with_spec(cellspec=spec_fcd)
            cell.set_cell_properties()
            
            h.celsius = 37

            ri = measure_input_resistance(cell, vinit)
            fr = measure_firing_rate(cell, stim_hz=input_hz, vinit=vinit, dt=0.1, weights=(82,36))

            ws.cell(row=ridx, column=4, value=ri)
            ws.cell(row=ridx, column=8, value=fr)
            del cell

            wb.save(xls_fname)  #save often to avoid sad accident 
        trial_from = 1
    inputhz_from = 0

