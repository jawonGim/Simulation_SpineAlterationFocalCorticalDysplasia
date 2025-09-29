from matplotlib import pyplot as plt

from openpyxl import Workbook, load_workbook
import os

def read_headers(ws):
    first_row = ws[1]

    print("HEADERS:")
    headers = []
    for cell in first_row :
        headers.append(cell.value)
        print(cell.value)
    return headers

def read_FI_data(fname = "../singleCell/Results/FI_for_various_dendcomb_2.xlsx") :
    from openpyxl import load_workbook

    wb = load_workbook(filename=fname)
    ws = wb.active

    headers = read_headers(ws)

    all_data = dict()
    for headername in headers :
        all_data[headername] = []

    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row) :
        col_vector = []
        for cell in col :
            col_vector.append(cell.value)
        all_data[col[0].value] = col_vector[1:]

    return all_data

def figureA_nbasl_ri(all_data) :
    # figure A, x: nbasal y: Ri 
    x = []
    y_max = []
    y_min = []
    for x_ in range(min(all_data["nbasal"]), max(all_data["nbasal"])+1) :
        y_max_ = 0
        y_min_ = 200
        for nbasal,ri  in zip(all_data["nbasal"], all_data["Ri_CTL"]) :
            if nbasal == x_ :
                if ri > y_max_ :
                    y_max_ = ri
                if ri < y_min_ :
                    y_min_ = ri

        x.append(x_)
        y_max.append(y_max_)
        y_min.append(y_min_)

    x.reverse()
    y_max.reverse()
    y_min.reverse()

    y_cen = []
    for ymax, ymin in zip(y_max, y_min) :
        yval = (ymax + ymin)/2
        y_cen.append(yval)

    plt.fill_between(x, y_min, y_max, alpha=.5, linewidth=0)
    plt.plot(x, y_cen, linewidth=1, color='black')

    plt.ylabel("Input Resistance (MΩ)")

    plt.xticks(ticks=[x[0], x[-1]], labels=["7,360", "7,090"]) 
    plt.xlabel("total dendritic length (µm)")
    plt.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=False,      # ticks along the bottom edge are off
        top=False)#,         # ticks along the top edge are off
    

    plt.show()  

def parsing_firingprob_data(ws) :
    print(f"max_row: {ws.max_row}, max_column: {ws.max_column}")
    num_activation = []
    for col in ws.iter_cols(min_col=1, max_col=1, min_row=1, max_row=ws.max_row) :
        for cell in col :
            num_activation.append(cell.value)
    
    avg_ = [] #hate this code..
    if ws.max_column == 23 : # nspines, 20 simulation results, sum of 20 simulations, stdev of 20 simulations
        for col in ws.iter_cols(min_col=ws.max_column-1, max_col=ws.max_column-1, min_row=1, max_row=ws.max_row) :
            for cell in col :
                avg_.append(cell.value/(ws.max_column-3))
    elif ws.max_column == 4 :   # nspines, number of simulations, sum of all simulations, stdev of all simulations
        for col in ws.iter_cols(min_col=ws.max_column-1, max_col=ws.max_column-1, min_row=1, max_row=ws.max_row) :
            for cell in col :
                avg_.append(cell.value/20) #for temp
    elif ws.max_column == 24 :  # nspines, 20 simulation results, sum, mean, stdev of 20 simulations
        for col in ws.iter_cols(min_col=ws.max_column-1, max_col=ws.max_column-1, min_row=1, max_row=ws.max_row) :
            for cell in col :
                avg_.append(cell.value)
    else:
        print(f"Error: ws.max_column is {ws.max_column}, not in considerations")
        return None
    

    error_ = []
    for col in ws.iter_cols(min_col=ws.max_column, max_col=ws.max_column, min_row=1, max_row=ws.max_row) :
        for cell in col :
            error_.append(cell.value)
                
    return num_activation, avg_, error_

def figureH_firingprob(fname="../singleCell/Results/merge_sum.xlsx", xpeak=[40, 175]) :
    from openpyxl import load_workbook

    print(fname)
    wb = load_workbook(filename=fname)

    # plt.figure()
    fig, (ax_fcd, ax_ctl) = plt.subplots(1, 2, sharey=True)
    fig.subplots_adjust(wspace=0.05)  # adjust space between Axes

    ws = wb['ctl']
    x, y, error_ = parsing_firingprob_data(ws)

    max_ = []
    min_ = []
    for v1, v2 in zip(y, error_):
        max_val = v1+v2
        if max_val > 1 :
            max_val = 1
        max_.append(max_val)
        min_val = v1-v2
        if min_val < 0 :
            min_val = 0
        min_.append(min_val)

    ax_ctl.fill_between(x, max_, min_, color='cyan', alpha=.55, linewidth=1)
    ax_ctl.plot(x, y, linewidth=2, color='darkcyan', label="Baseline")
    ax_fcd.plot(x, y, linewidth=2, color='darkcyan', label="Baseline")
    
    
    ws = wb['fcd']
    x_, y, error_ = parsing_firingprob_data(ws)

    max_ = []
    min_ = []
    for v1, v2 in zip(y, error_):
        max_val = v1+v2
        if max_val > 1 :
            max_val = 1
        max_.append(max_val)
        min_val = v1-v2
        if min_val < 0 :
            min_val = 0
        min_.append(min_val)

    ax_fcd.fill_between(x, max_, min_, color='magenta', alpha=.55, linewidth=1)
    ax_fcd.plot(x, y, linewidth=2, color='darkmagenta', label="Altered")
    ax_ctl.plot(x, y, linewidth=2, color='darkmagenta', label="Altered")  #just for the legend

    
    # xpeak = [40,175]
    ax_fcd.set_xlim(xpeak[0]-10, xpeak[0]+10)  
    ax_ctl.set_xlim(xpeak[1]-10, xpeak[1]+10)  

    ax_fcd.spines.right.set_visible(False)
    ax_ctl.spines.left.set_visible(False)
    ax_fcd.yaxis.tick_left()
    ax_ctl.yaxis.tick_right()

    ax_fcd.set_xticks([xpeak[0]-5, xpeak[0]+5], labels=[str(xpeak[0]-5), str(xpeak[0]+5)], fontsize=12) 
    ax_ctl.set_xticks([xpeak[1]-5, xpeak[1]+5], labels=[str(xpeak[1]-5), str(xpeak[1]+5)], fontsize=12) 

    ax_fcd.set_yticks([0, 1.0], labels=["0.0", "1.0"], fontsize=12)
    
    d = .2  # proportion of vertical to horizontal extent of the slanted line
    kwargs = dict(marker=[(-d, -1), (d, 1)], markersize=10,
                linestyle="none", color='k', mec='k', mew=1, clip_on=False)
    ax_fcd.plot([1, 1], [0, 1], transform=ax_fcd.transAxes, **kwargs)
    ax_ctl.plot([0, 0], [0, 1], transform=ax_ctl.transAxes, **kwargs)

    ax_fcd.set_ylabel("Firing Probability", fontsize=14)
    
    ax_fcd.set_xlabel('Number of Activated Spines', fontsize=14)
    ax_fcd.xaxis.set_label_coords(0.5, 0.05, transform=fig.transFigure)

    plt.savefig(f"syn_activation.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    plt.show()  

def figureJ_FI_new(all_data) :
    plt.figure()
    # figure c, x: input_hz  y: firing rate
    x = []
    fr_max_ctl = []
    fr_min_ctl = []

    fr_max_fcd = []
    fr_min_fcd = []

    fr_avg_ctl = []
    fr_avg_fcd = []

    fr_ctl_nbasalA = []
    fr_ctl_nbasalB = []
    fr_fcd_nbasalA = []
    fr_fcd_nbasalB = []

    for hz_ in range(0, 25) :
        input_hz = 0.2 + 0.2*hz_
        input_hz = round(input_hz, 1)

        x.append(input_hz)

        fr_max_ctl_ = 0
        fr_min_ctl_ = 500
        fr_max_fcd_ = 0
        fr_min_fcd_ = 500

        # get all indices of the same input_hz
        idx_ihz = [i for i, elem in enumerate(all_data["I_hz"]) if round(elem, 1) == input_hz]
        fr_ctl_Ihz = [all_data["fr_CTL"][i] for i in idx_ihz]
        fr_fcd_Ihz = [all_data["fr_FCD"][i] for i in idx_ihz]

        # get all indices of the same input_hz and specific number of basal dendrites
        idx_ihz_nbasalA = [i for i, elem in enumerate(all_data["I_hz"]) if round(elem, 1) == input_hz and all_data["nbasal"][i] == 2]
        idx_ihz_nbasalB = [i for i, elem in enumerate(all_data["I_hz"]) if round(elem, 1) == input_hz and all_data["nbasal"][i] == 8]

        fr_ctl_Ihz_nbasalA = [all_data["fr_CTL"][i] for i in idx_ihz_nbasalA]
        fr_ctl_Ihz_nbasalB = [all_data["fr_CTL"][i] for i in idx_ihz_nbasalB]
        fr_fcd_Ihz_nbasalA = [all_data["fr_FCD"][i] for i in idx_ihz_nbasalA]
        fr_fcd_Ihz_nbasalB = [all_data["fr_FCD"][i] for i in idx_ihz_nbasalB]

        # get and append all results values 
        fr_max_ctl.append(max(fr_ctl_Ihz))
        fr_min_ctl.append(min(fr_ctl_Ihz))
        fr_max_fcd.append(max(fr_fcd_Ihz))
        fr_min_fcd.append(min(fr_fcd_Ihz))
        fr_avg_ctl.append(sum(fr_ctl_Ihz)/len(fr_ctl_Ihz))
        fr_avg_fcd.append(sum(fr_fcd_Ihz)/len(fr_fcd_Ihz))

        fr_ctl_nbasalA.append(sum(fr_ctl_Ihz_nbasalA)/len(fr_ctl_Ihz_nbasalA))
        fr_ctl_nbasalB.append(sum(fr_ctl_Ihz_nbasalB)/len(fr_ctl_Ihz_nbasalB))
        fr_fcd_nbasalA.append(sum(fr_fcd_Ihz_nbasalA)/len(fr_fcd_Ihz_nbasalA))
        fr_fcd_nbasalB.append(sum(fr_fcd_Ihz_nbasalB)/len(fr_fcd_Ihz_nbasalB))


    plt.fill_between(x, fr_min_ctl, fr_max_ctl, color='cyan', alpha=.55, linewidth=1)
    plt.plot(x, fr_avg_ctl, linewidth=1.5, color='darkcyan', label="Baseline")

    plt.fill_between(x, fr_min_fcd, fr_max_fcd, color='magenta', alpha=.55, linewidth=1)
    plt.plot(x, fr_avg_fcd, linewidth=1.5, color='darkmagenta', label="Altered")

    plt.ylabel("Firing Rate (Hz)", fontsize=14)
    plt.ylim((0, 170))
    plt.xlim((0.2,5))
    plt.yticks(ticks=[0, 50, 100, 150], labels=["0", "50", "100", "150"], fontsize=12)
    plt.xlabel("Input Frequency (Hz)", fontsize=14)
    plt.xticks(ticks=[0.2, 1.0, 2.0, 3.0, 4.0, 5.0], labels=["0.2", "1.0", "2.0", "3.0", "4.0", "5.0"], fontsize=12)
    plt.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=True,      # ticks along the bottom edge are off
        top=False)#,         # ticks along the top edge are off

    plt.savefig(f"FI_curve.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    plt.show()  

# distance from soma vs somatic amplitude 
# scatter plot and fitting line
# does not include dendrite information
def distance_from_soma_vs_somatic_amplitude(N=100, show=True, legend=False) :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell
    import numpy as np

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 200)/h.dt

    amplitude_ctl = []
    dist_spine_ctl = []

    for i in range(1, N+1) :
        cellA = cell.cell(name="control", gid=i)
        cellA.create_cell()

        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        cellA.set_cell_properties(sf=2.0)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amplitude_ctl.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))
        dist_spine_ctl.append(h.distance(cellA.soma(1), cellA.spine_necks[0](0)))

        del cellA


    amplitude_fcd = []
    dist_spine_fcd = []
    for i in range(1, N+1) :
        cellA = cell.cell(name="epileptogenic", gid=i)
        cellA.create_cell()
        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        cellA.set_cell_properties(sf=1.2)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amplitude_fcd.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))
        dist_spine_fcd.append(h.distance(cellA.soma(1), cellA.spine_necks[0](0)))

        del cellA


    #sort
    dist_spine_ctl, amplitude_ctl = zip(*sorted(zip(dist_spine_ctl, amplitude_ctl)))
    dist_spine_fcd, amplitude_fcd = zip(*sorted(zip(dist_spine_fcd, amplitude_fcd)))


    fig, ax = plt.subplots()
    
    plt.scatter(dist_spine_ctl, amplitude_ctl, marker="o", s=20, color="cyan", alpha=0.7, label="Baseline")
    plt.scatter(dist_spine_fcd, amplitude_fcd, marker="o", s=20, color="magenta", alpha=0.7, label="Altered")


    # Plot the scatter plot and the polynomial trend line
    coefficients = np.polyfit(dist_spine_ctl, amplitude_ctl, 2)
    p = np.poly1d(coefficients)    
    plt.plot(dist_spine_ctl, p(dist_spine_ctl), linewidth=2, color='darkcyan', label='Baseline')

    coefficients = np.polyfit(dist_spine_fcd, amplitude_fcd, 2)
    p = np.poly1d(coefficients)    
    plt.plot(dist_spine_fcd, p(dist_spine_fcd), linewidth=2, color='darkmagenta', label='Altered')


    plt.xlabel("Spine Location Relative to Soma (µm)", fontsize=14)
    plt.ylabel("Somatic EPSP Amplitude (mV)", fontsize=14)

    
    plt.xlim((30, 390))
    plt.ylim((0.1, 0.42))

    ax.set_xticks(ticks=[100, 200, 300], labels=["100", "200", "300"], fontsize=12)
    ax.set_yticks(ticks=[0.1, 0.2, 0.3, 0.4], labels=["0.1", "0.2", "0.3", "0.4"], fontsize=12)
    plt.tick_params(
        axis='y',          # changes apply to the y-axis
        which='both',      # both major and minor ticks are affected
        right=False 
        )
    if legend :
        plt.legend(loc="upper right")

    # Force all spines on
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color('black')

    plt.savefig(f"distance_vs_soma_amp.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    if show :
        plt.show()

# fill between plot about distance from soma vs somatic amplitude
# without AVG line 
# upper line from basal dendrite, lower line from apical dendrite 
def distance_from_soma_vs_somatic_amplitude_(N=100, show=True, legend=False) :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 200)/h.dt

    amplitude_ctl = []
    dist_spine_ctl = []
    dend_type_ctl = []
    for i in range(1, N+1) :
        cellA = cell.cell(name="control")
        cellA.create_cell()

        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        cellA.set_cell_properties(sf=2.0)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amplitude_ctl.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))
        dist_spine_ctl.append(h.distance(cellA.soma(1), cellA.spine_necks[0](0)))
        dend_type_ctl.append(cellA.spines_on)

        del cellA


    amplitude_fcd = []
    dist_spine_fcd = []
    dend_type_fcd = []
    for i in range(1, N+1) :
        cellA = cell.cell(name="epileptogenic")
        cellA.create_cell()
        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        cellA.set_cell_properties(sf=1.2)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amplitude_fcd.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))
        dist_spine_fcd.append(h.distance(cellA.soma(1), cellA.spine_necks[0](0)))
        dend_type_fcd.append(cellA.spines_on)

        del cellA


    fig, ax = plt.subplots()

    #apical
    idx = [i for i, x in enumerate(dend_type_ctl) if x == "A"]
    distance_apical_ctl = [dist_spine_ctl[i] for i in idx]
    amplitude_apical_ctl = [amplitude_ctl[i] for i in idx]

    idx = [i for i, x in enumerate(dend_type_fcd) if x == "A"]
    distance_apical_fcd = [dist_spine_fcd[i] for i in idx]
    amplitude_apical_fcd = [amplitude_fcd[i] for i in idx]

    #basal
    idx = [i for i, x in enumerate(dend_type_ctl) if x == "B"]
    distance_basal_ctl = [dist_spine_ctl[i] for i in idx]
    amplitude_basal_ctl = [amplitude_ctl[i] for i in idx]

    idx = [i for i, x in enumerate(dend_type_fcd) if x == "B"]
    distance_basal_fcd = [dist_spine_fcd[i] for i in idx]
    amplitude_basal_fcd = [amplitude_fcd[i] for i in idx]

    #sort by distance
    zip_ = sorted(zip(distance_apical_ctl, amplitude_apical_ctl))
    distance_apical_ctl, amplitude_apical_ctl = zip(*zip_)

    zip_ = sorted(zip(distance_basal_ctl, amplitude_basal_ctl))
    distance_basal_ctl, amplitude_basal_ctl = zip(*zip_)

    zip_ = sorted(zip(distance_apical_fcd, amplitude_apical_fcd))
    distance_apical_fcd, amplitude_apical_fcd = zip(*zip_)

    zip_ = sorted(zip(distance_basal_fcd, amplitude_basal_fcd))
    distance_basal_fcd, amplitude_basal_fcd = zip(*zip_)

    import numpy as np
    x_comb_ctl = np.unique(np.concatenate((np.array(distance_apical_ctl), np.array(distance_basal_ctl))))
    x_comb_fcd = np.unique(np.concatenate((np.array(distance_apical_fcd), np.array(distance_basal_fcd))))

    amp_actl = np.interp(x_comb_ctl, distance_apical_ctl, amplitude_apical_ctl) 
    amp_bctl = np.interp(x_comb_ctl, distance_basal_ctl, amplitude_basal_ctl)

    amp_afcd = np.interp(x_comb_fcd, distance_apical_fcd, amplitude_apical_fcd)
    amp_bfcd = np.interp(x_comb_fcd, distance_basal_fcd, amplitude_basal_fcd)

    
    plt.fill_between(x_comb_ctl, amp_actl, amp_bctl, color="cyan", alpha=0.7, label="Baseline")
    plt.fill_between(x_comb_fcd, amp_afcd, amp_bfcd, color="magenta", alpha=0.7, label="Altered")


    plt.xlabel("Distance from Soma (µm)")
    plt.ylabel("Somatic EPSP Amplitude (mV)")

    
    plt.xlim((40, 400))

    ax.set_xticks(ticks=[100, 200, 300, 400], labels=["100", "200", "300", "400"])
    ax.set_yticks(ticks=[0.1, 0.2, 0.3, 0.4], labels=["0.1", "0.2", "0.3", "0.4"])
    plt.tick_params(
        axis='y',          # changes apply to the y-axis
        which='both',      # both major and minor ticks are affected
        right=False 
        )
    if legend :
        plt.legend(loc="upper right")

    plt.savefig(f"distance_vs_somaticAmp.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    if show :
        plt.show()

# neck diameter vs delta V (head-base)
# fill between plot and marker for each group
def neck_diam_vs_deltaV_head_base(N=20, show=True, legend=False) :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 50)/h.dt

    x_ = []
    y_max = []
    y_min = []
    for neck_factor in range(0, 100) :
        neck_diam = round(0.05 + neck_factor*0.01, 2)
        if neck_diam > 0.52: #1.0:
            break
        x_.append(neck_diam)
        diff_ = []
        for i in range(1, N+1) :
            cellA = cell.cell(name="control")
            cellspec = dict()
            cellspec["spine_neck_diam"] = neck_diam
            cellA.create_cell_with_spec(cellspec)

            ampa_ctl = h.NetCon(stim, cellA.ampar[0])
            ampa_ctl.weight[0] = 31     #ampa_weight
            ampa_ctl.delay = 0 * ms
            nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
            nmda_ctl.weight[0] = 14     #nmda_weight
            nmda_ctl.delay = 0 * ms

            tstop = 1_000 * ms

            t = h.Vector().record(h._ref_t)
            vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
            vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

            cellA.set_cell_properties(sf=2.0)

            h.finitialize(vinit)
            h.continuerun(tstop)      

            diff_.append(vhead.max(compute_start, compute_end) - vbase.max(compute_start, compute_end))

            del cellA
        y_max.append(max(diff_))
        y_min.append(min(diff_))

    fig, ax = plt.subplots()

    plt.fill_between(x_, y_min, y_max, color="lightgrey", alpha=0.9)

    plt.xlabel("Spine Neck Diameter (µm)", fontsize=14)
    plt.ylabel('ΔV (Head - Base, mV)', fontsize=14)
    plt.ylim((0, max(y_max)))
    plt.xlim((0,0.52))
    plt.yticks(ticks=[0, 5, 10, 15], labels=["0", "5", "10", "15"], fontsize=12)
    plt.xticks(ticks=[0.1, 0.2, 0.3, 0.4, 0.5], labels=["0.1", "0.2", "0.3", "0.4", "0.5"], fontsize=12)

    plt.axvline(x = 0.15, ymin = 0, ymax = y_max[x_.index(0.15)]/max(y_max), color = 'cyan', label = 'Baseline', linestyle='-')
    plt.axvline(x = 0.3, ymin = 0, ymax = y_max[x_.index(0.3)]/max(y_max), color = 'magenta', label = 'Altered', linestyle='-')

    plt.scatter(0.3, y_max[x_.index(0.3)], s=20, c='magenta', marker='X')
    plt.scatter(0.15, y_max[x_.index(0.15)], s=20, c='cyan', marker='X')
    if legend :
        plt.text(0.3, y_max[x_.index(0.3)], 'Altered', fontdict={'size': 9, 'color': 'magenta'})
        plt.text(0.15, y_max[x_.index(0.15)], 'Baseline', fontdict={'size': 9, 'color': 'cyan'})


    # Force all spines on
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color('black')

    plt.savefig(f"neckdiam_deltaV.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    if show :
        plt.show()
 

# delta V accumulation plot when neck diameter changed
def deltaV_by_neckdiam_accumulative(N=20, show=True, legend=False) :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 50)/h.dt

    x_ = []
    y_max = []
    y_min = []
    diams = [0.150, 0.300]
    color_ = ["cyan", "magenta"]
    for neck_diam in diams :
        diff_ = []
        for i in range(1, N+1) :
            cellA = cell.cell(name="control")
            cellspec = dict()
            cellspec["spine_neck_diam"] = neck_diam
            cellA.create_cell_with_spec(cellspec)

            ampa_ctl = h.NetCon(stim, cellA.ampar[0])
            ampa_ctl.weight[0] = 31     #ampa_weight
            ampa_ctl.delay = 0 * ms
            nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
            nmda_ctl.weight[0] = 14     #nmda_weight
            nmda_ctl.delay = 0 * ms

            tstop = 1_000 * ms

            t = h.Vector().record(h._ref_t)
            vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
            vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

            cellA.set_cell_properties(sf=2.0)

            h.finitialize(vinit)
            h.continuerun(tstop)      

            diff_.append(vhead.max(compute_start, compute_end) - vbase.max(compute_start, compute_end))
            # diff_.append((vhead.max(compute_start, compute_end) - vbase.max(compute_start, compute_end))/vbase.max(compute_start, compute_end))

            del cellA

        plt.hist(diff_, bins=100, cumulative=True, density=True, histtype='stepfilled', color=color_[diams.index(neck_diam)], alpha=0.5)


    if show :
        plt.show()
 

# scatter plot of neck diameter vs delta V (head-base)
def neck_diam_vs_differ_head_base_scatter_() :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell
    import random

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 300)/h.dt

    amp_head_ctl = []
    amp_head_fcd = []
    amp_delta_spine_ctl = []
    amp_delta_spine_fcd = []

    N = 100
    for i in range(1, N+1) :
        cellA = cell.cell(name="control")
        cellA.create_cell()

        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
        vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

        cellA.set_cell_properties(sf=2.0)  

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amp_head_ctl.append(vhead.max(compute_start, compute_end) - vhead.min(compute_start, compute_end))
        amp_delta_spine_ctl.append(vhead.max(compute_start, compute_end) - vbase.max(compute_start, compute_end))

        del cellA


        cellB = cell.cell(name="epileptogenic")
        # cellB.create_cell()
        cellspec = dict()
        cellspec["spine_neck_diam"] = 0.301
        cellB.create_cell_with_spec(cellspec)

        ampa_fcd = h.NetCon(stim, cellB.ampar[0])
        ampa_fcd.weight[0] = 31     #ampa_weight
        ampa_fcd.delay = 0 * ms
        nmda_fcd = h.NetCon(stim, cellB.nmdar[0])
        nmda_fcd.weight[0] = 14     #nmda_weight
        nmda_fcd.delay = 0 * ms


        vhead_ = h.Vector().record(cellB.spine_heads[0](0.5)._ref_v)
        vbase_ = h.Vector().record(cellB.spine_loc_dend(cellB.spine_loc_in_dends)._ref_v)

        cellB.set_cell_properties(sf=2.0)  

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amp_head_fcd.append(vhead_.max(compute_start, compute_end) - vhead_.min(compute_start, compute_end))
        amp_delta_spine_fcd.append(vhead_.max(compute_start, compute_end) - vbase_.max(compute_start, compute_end))

        del cellB

    #plot 그리기
    fig, ax = plt.subplots()

    plt.scatter(amp_head_ctl, amp_delta_spine_ctl, color="cyan", s=5, label="Control")
    plt.scatter(amp_head_fcd, amp_delta_spine_fcd, color="magenta", s=5, marker="*", label="Epileptogenic")
    plt.xlabel("Head EPSP Amplitude (mV)")
    plt.ylabel("ΔV (Head - Base, mV)")
    plt.legend(loc="upper left")
    
    plt.xlim((1, 9))
    plt.ylim((0, 4))

    plt.xticks(ticks=[2, 4, 6, 8], labels=["2", "4", "6", "8"])
    plt.yticks(ticks=[0, 1, 2, 3], labels=["0", "1", "2", "3"])

    plt.savefig(f"scatter_dvhead_dvspine.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)
    plt.show()

# head diameter(fixed neck length) vs head, base, soma EPSP amplitude 
def head_diam_vs_amplitude_multipoints(N=20) :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 300)/h.dt


    fig, (ax_spine, ax_soma) = plt.subplots(2, 1, sharex=True, gridspec_kw={'height_ratios': [0.7, 0.3]})

    
    fig.subplots_adjust(hspace=0.05)  # adjust space between Axes

    head_ = []
    base_ =[]
    soma_ = []
    head_max = []
    base_max = []
    soma_max = []
    head_min = []
    base_min = []
    soma_min = []

    cellspec = dict()
    diam_ticks = 20*2 
    diam_term = 0.05
    x_ = []
    tick = 0.01
    head_diam = 0.2     #start
    while True :
        amp_head_ctl = []
        amp_base_ctl = []
        amp_soma_ctl = []


        x_.append(head_diam)
        for i in range(1, N+1) :
            cellA = cell.cell(name="control")
            cellspec["spine_head_diam"] = head_diam
            cellspec["spine_len"] = 1.3 + head_diam #1.3 is for neck length
            cellA.create_cell_with_spec(cellspec)
            

            ampa_ctl = h.NetCon(stim, cellA.ampar[0])
            ampa_ctl.weight[0] = 31     #ampa_weight
            ampa_ctl.delay = 0 * ms
            nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
            nmda_ctl.weight[0] = 14     #nmda_weight
            nmda_ctl.delay = 0 * ms

            tstop = 1_000 * ms

            vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
            vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

            cellA.set_cell_properties(sf=2.0)  

            h.finitialize(vinit)
            h.continuerun(tstop)      

            amp_head_ctl.append(vhead.max(compute_start, compute_end) - vhead.min(compute_start, compute_end))
            amp_base_ctl.append(vbase.max(compute_start, compute_end) - vbase.min(compute_start, compute_end))
            amp_soma_ctl.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))

            del cellA

        head_.append(sum(amp_head_ctl)/N)
        base_.append(sum(amp_base_ctl)/N)
        soma_.append(sum(amp_soma_ctl)/N)
        head_max.append(max(amp_head_ctl))
        base_max.append(max(amp_base_ctl))
        soma_max.append(max(amp_soma_ctl))
        head_min.append(min(amp_head_ctl))
        base_min.append(min(amp_base_ctl))
        soma_min.append(min(amp_soma_ctl))

        head_diam = head_diam + 0.01
        head_diam = round(head_diam, 2)
        if head_diam > 1.5 :#2.0 :
            break
    

    ax_spine.plot(x_, head_, color="red",label="Head")
    ax_spine.plot(x_, base_, color="blue", label="Base")
    ax_spine.plot(x_, soma_, color="black", label="Soma")

    ax_soma.plot(x_, head_, color="red",label="Head")
    ax_soma.plot(x_, base_, color="blue", label="Base")
    ax_soma.plot(x_, soma_, color="black", label="Soma")

    ax_spine.fill_between(x_, head_min, head_max, color="red", alpha=0.3)
    ax_spine.fill_between(x_, base_min, base_max, color="blue", alpha=0.3)
    ax_spine.fill_between(x_, soma_min, soma_max, color="black", alpha=0.3)

    ax_soma.fill_between(x_, head_min, head_max, color="red", alpha=0.3)
    ax_soma.fill_between(x_, base_min, base_max, color="blue", alpha=0.3)
    ax_soma.fill_between(x_, soma_min, soma_max, color="black", alpha=0.3)

    
    # ax_spine.xaxis.tick_top()
    ax_spine.spines.bottom.set_visible(False)

    ax_soma.xaxis.tick_bottom()
    ax_soma.spines.top.set_visible(False)

    ax_spine.set_xticks(ticks=[])
    ax_spine.set_yticks(ticks=[3,5,7], labels=["3", "5", "7"])
    ax_spine.set_ylim([1,9])
    

    ax_soma.set_xticks(ticks=[0.2, 0.5, 1.0, 1.5], labels=["0.2", "0.5", "1.0", "1.5"], fontsize=12)
    ax_soma.set_yticks(ticks=[0, 0.4], labels=["0", "0.4"], fontsize=12)
    ax_soma.set_ylim([0, 0.5])
    
    ax_spine.axvline(x = 0.55, ymin = 0, ymax = 1, color = 'cyan', linestyle='-')
    ax_spine.axvline(x = 0.70, ymin = 0, ymax = 1, color = 'magenta',linestyle='-')
    ax_soma.axvline(x = 0.55, ymin = 0, ymax = 1, color = 'cyan', linestyle='-')
    ax_soma.axvline(x = 0.70, ymin = 0, ymax = 1, color = 'magenta', linestyle='-')


    d = .5 # proportion of vertical to horizontal extent of the slanted line
    kwargs = dict(marker=[(-d, -1), (d, 1)], markersize=10,
                linestyle="none", color='k', mec='k', mew=1, clip_on=False)
    ax_spine.plot([0, 1], [0, 0], transform=ax_spine.transAxes, **kwargs)
    ax_soma.plot([0, 1], [1, 1], transform=ax_soma.transAxes, **kwargs)

    ax_soma.set_xlabel("Spine Head Diameter (µm)", fontsize=14)
    ax_spine.legend(loc="upper center", ncol=3, fontsize=12)


    fig.text(0.04, 0.5, 'EPSP Amplitude (mV)', va='center', rotation='vertical', fontsize=14)

    plt.savefig(f"head_diam_amplitude.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)
    plt.show()

# violin plot for head, base, soma EPSP amplitude 
# and saving the raw values as xlsx file
def violin_and_rawdata(alter, N=100, show=True, legend=False) :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 300)/h.dt

    amp_head_ctl = []
    amp_head_fcd = []
    amp_base_ctl = []
    amp_base_fcd = []
    amp_soma_ctl = []
    amp_soma_fcd = []

    for i in range(1, N+1) :
        cellA = cell.cell(name="control")
        cellA.create_cell()

        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        vhead = h.Vector().record(cellA.spine_heads[0](0.5)._ref_v)
        vbase = h.Vector().record(cellA.spine_loc_dend(cellA.spine_loc_in_dends)._ref_v)

        cellA.set_cell_properties(sf=2.0)  

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amp_head_ctl.append(vhead.max(compute_start, compute_end) - vhead.min(compute_start, compute_end))
        amp_base_ctl.append(vbase.max(compute_start, compute_end) - vbase.min(compute_start, compute_end))
        amp_soma_ctl.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))

        del cellA


        cellB = cell.cell(name="epileptogenic")
        cellspec = dict()
        cellspec["model"] = "epilepsy"

        if alter == "neck" :
            cellspec["spine_neck_diam"] = 0.301
        elif alter == "head" :
            cellspec["spine_head_diam"] = 0.718
            cellspec["spine_len"] = 1.823-0.551+0.718
        else :
            pass
    
        cellB.create_cell_with_spec(cellspec)

        ampa_fcd = h.NetCon(stim, cellB.ampar[0])
        ampa_fcd.weight[0] = 31     #ampa_weight
        ampa_fcd.delay = 0 * ms
        nmda_fcd = h.NetCon(stim, cellB.nmdar[0])
        nmda_fcd.weight[0] = 14     #nmda_weight
        nmda_fcd.delay = 0 * ms


        vhead_ = h.Vector().record(cellB.spine_heads[0](0.5)._ref_v)
        vbase_ = h.Vector().record(cellB.spine_loc_dend(cellB.spine_loc_in_dends)._ref_v)

        if alter == "density" :
            cellB.set_cell_properties(sf=1.2)
        else :
            cellB.set_cell_properties(sf=2.0)  

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amp_head_fcd.append(vhead_.max(compute_start, compute_end) - vhead_.min(compute_start, compute_end))
        amp_base_fcd.append(vbase_.max(compute_start, compute_end) - vbase_.min(compute_start, compute_end))
        amp_soma_fcd.append(cellB.v_soma.max(compute_start, compute_end) - cellB.v_soma.min(compute_start, compute_end))

        del cellB

    #plot 그리기
    fig, ax = plt.subplots(figsize=(4,4))

    # #Head, Base, Soma position base
    box_colors = ["cyan", "magenta", "cyan", "magenta", "cyan", "magenta"]
    bplot1 = ax.violinplot([amp_head_ctl, amp_head_fcd], 
                           positions=[1,2], widths=0.7, showmedians=True)
    bplot3 = ax.violinplot([amp_base_ctl, amp_base_fcd], 
                           positions=[3,4], widths=0.7, showmedians=True)

    bplot1['cmaxes'].set_color("black")
    bplot1['cmins'].set_color("black")
    bplot1['cmedians'].set_color('black') 

    bplot3['cmaxes'].set_color("black")
    bplot3['cmins'].set_color("black")  
    bplot3['cmedians'].set_color('black') 
    
    ax.vlines(1, min(amp_head_ctl), max(amp_head_ctl), color="black", linestyle='-', lw=1.5)
    ax.vlines(2, min(amp_head_fcd), max(amp_head_fcd), color="black", linestyle='-', lw=1.5)
    ax.vlines(3, min(amp_base_ctl), max(amp_base_ctl), color="black", linestyle='-', lw=1.5)
    ax.vlines(4, min(amp_base_fcd), max(amp_base_fcd), color="black", linestyle='-', lw=1.5)
    for pc, color in zip(bplot1['bodies'], box_colors[0:2]):
        pc.set_facecolor(color)
        pc.set_edgecolor(color)
        pc.set_alpha(0.5)
    for pc, color in zip(bplot3['bodies'], box_colors[2:4]):
        pc.set_facecolor(color)
        pc.set_edgecolor(color)
        pc.set_alpha(0.5)
        
    if legend :
        plt.legend(["Baseline","Altered"], loc="lower left")

    ax.set_xticks([1.5, 3.5])
    ax.set_xticklabels(["Head", "Base"], fontsize=12)
    ax.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=False,      # ticks along the bottom edge are off
        top=False)

    ax.set_ylabel('EPSP Amplitude (mV)', fontsize=14)
    ax.set_yticks(ticks=[0, 2, 4, 6, 8, 10], labels=["0", "2", "4", "6", "8", "10"], fontsize=12)
    ax.set_ylim([0, 10.5])
    ax.set_title('Spine', fontsize=14)
    
    # Force all spines on
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color('black')

    plt.savefig(f"violin_{alter}_ampl_spine.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    fig, ax2 = plt.subplots(figsize=(2,4))
    bplot2 = ax2.violinplot([amp_soma_ctl, amp_soma_fcd], positions=[5,6], widths=0.7, showmedians=True)
    bplot2['cmaxes'].set_color("black")
    bplot2['cmins'].set_color("black")
    bplot2['cmedians'].set_color('black')
    ax2.vlines(5, min(amp_soma_ctl), max(amp_soma_ctl), color="black", linestyle='-', lw=1.5)
    ax2.vlines(6, min(amp_soma_fcd), max(amp_soma_fcd), color="black", linestyle='-', lw=1.5)
    for pc, color in zip(bplot2['bodies'], box_colors[4:6]):
        pc.set_facecolor(color)
        pc.set_edgecolor(color)
        pc.set_alpha(0.5)


    ax2.set_yticks(ticks=[0, 0.2, 0.4], labels=["0", "0.2", "0.4"], fontsize=12)
    ax2.set_xticks([])
    ax2.set_ylim([0, 0.42])
    ax2.tick_params(top=True, labeltop=True, bottom=False, labelbottom=False)
    ax2.set_title('Soma', fontsize=14)
    # Set the y-axis label position to the right
    ax2.yaxis.set_label_position("right")
    ax2.yaxis.tick_right()


    # Force all spines on
    for spine in ax2.spines.values():
        spine.set_visible(True)
        spine.set_color('black')

    plt.savefig(f"violin_{alter}_ampl_soma.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    if show :
        plt.show()

    xls_fname = f"violin_{alter}_rawdata.xlsx"
    if os.path.isfile(xls_fname) :
        wb = load_workbook(filename=xls_fname)
    else :
        wb = Workbook()
    ws = wb.active


    ws.cell(row=1, column=1, value="head_ctl")
    ws.cell(row=1, column=2, value="head_fcd")
    ws.cell(row=1, column=3, value="base_ctl")
    ws.cell(row=1, column=4, value="base_fcd")
    ws.cell(row=1, column=5, value="soma_ctl")
    ws.cell(row=1, column=6, value="soma_fcd")

    for irow in range(1, N+1) :
        ws.cell(row=irow+1, column=1, value=amp_head_ctl[irow-1])
        ws.cell(row=irow+1, column=2, value=amp_head_fcd[irow-1])
        ws.cell(row=irow+1, column=3, value=amp_base_ctl[irow-1])
        ws.cell(row=irow+1, column=4, value=amp_base_fcd[irow-1])
        ws.cell(row=irow+1, column=5, value=amp_soma_ctl[irow-1])
        ws.cell(row=irow+1, column=6, value=amp_soma_fcd[irow-1])

    wb.save(xls_fname)

# spine density vs input resistance
def density_vs_RI(show=True, legend=False) :
    from neuron import h
    from neuron.units import ms, mV

    # import cell_singlespine as cellA
    import cell_singlespine_randomloc as cellA

    import sys
    sys.path.append("../")

    from singleCell.utils import measure_input_resistance

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    x = []
    y = []
    for sf_ in range(0, 100) :
        sf = round(0.8 + sf_*0.05, 2)
        if sf>2.5 :
            break

        x.append(sf)

        cell_ = cellA.cell(name="control")
        cell_.create_cell()

        ampa_ctl = h.NetCon(stim, cell_.ampar[0])
        ampa_ctl.weight[0] = 31 #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cell_.nmdar[0])
        nmda_ctl.weight[0] = 14 #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        #mesure RI with varius factor 
        cell_.set_cell_properties(sf=sf)
        ri = measure_input_resistance(cell_)
        y.append(ri)


        h.finitialize(vinit)
        h.continuerun(tstop)

        del cell_

    fig, ax = plt.subplots()
    plt.plot(x, y, color="black", linewidth=2.5)

    plt.ylim((70, 230))
    # Set the y-axis label position to the right
    ax.yaxis.set_label_position("right")

    # Move the y-axis ticks and tick labels to the right
    ax.yaxis.tick_right()

    ax.tick_params(
            axis='y',          # changes apply to the y-axis
            which='both',      # both major and minor ticks are affected
            left=False,      # ticks along the bottom edge are off
            right=True)
    plt.yticks(ticks=[100, 150, 200], labels=["100", "150", "200"], fontsize=12)

    plt.xticks(ticks=[x[0], x[-1]], labels=["Sparse", "Dense"], fontsize=12)
    plt.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=True,      # ticks along the bottom edge are off
        top=False)#),         # ticks along the top edge are off

    # Force all spines on
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color('black')
    
    #flip X axis 
    plt.gca().invert_xaxis()
        
    plt.scatter(1.2, y[x.index(1.20)], s=120, c='magenta', marker='X')
    plt.scatter(2.0, y[x.index(2.00)], s=120, c='cyan', marker='X')

    if legend :
        plt.text(2.0, y[x.index(2.00)], 'Baseline', fontdict={'size': 9, 'color': 'cyan'})
        plt.text(1.2, y[x.index(1.20)], 'Altered', fontdict={'size': 9, 'color': 'magenta'})

    plt.xlabel("Spine Density", fontsize=14)
    plt.ylabel("Input Resistance (MΩ)", fontsize=14)
    plt.savefig("density_RI.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    if show :
        plt.show()

# plot firing probability vs number_ of activated spines as single plot (not broken axis)
def figureH_firingprob_singleplot(fname="../singleCell/Results/merge_sum.xlsx", x_range=[0,40], tick_=10) :
    from openpyxl import load_workbook

    print(fname)
    wb = load_workbook(filename=fname)

    plt.figure()

    ws = wb['ctl']
    x, y, error_ = parsing_firingprob_data(ws)

    max_ = []
    min_ = []
    for v1, v2 in zip(y, error_):
        max_val = v1+v2
        if max_val > 1 :
            max_val = 1
        max_.append(max_val)
        min_val = v1-v2
        if min_val < 0 :
            min_val = 0
        min_.append(min_val)

    plt.plot(x, y, linewidth=1, color='cyan', label="Baseline")
    
    ws = wb['fcd']
    x_, y, error_ = parsing_firingprob_data(ws)

    max_ = []
    min_ = []
    for v1, v2 in zip(y, error_):
        max_val = v1+v2
        if max_val > 1 :
            max_val = 1
        max_.append(max_val)
        min_val = v1-v2
        if min_val < 0 :
            min_val = 0
        min_.append(min_val)

    plt.plot(x_, y, linewidth=1, color='magenta', label="Altered")
    
    plt.xlim(x_range)  

    xticks_ = [i for i in range(x_range[0], x_range[1]+1, tick_)]
    xticklabels_ = [str(i) for i in xticks_]
    plt.xticks(xticks_,  labels=xticklabels_)
    
    plt.ylabel("Firing Probability", fontsize=14)
    plt.xlabel('Number of Activated Spines', fontsize=14)

    plt.show()

## test function for the spine distance vs somatic EPSP amplitude
def scatter_distance_somaticEPSP(N=100, show=True) :
    from neuron import h
    from neuron.units import mV, ms
    import cell_singlespine_randomloc as cell
    import numpy as np

    h.load_file("stdrun.hoc")

    vinit = -70 * mV

    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    compute_start = (stim.start - 10)/h.dt
    compute_end = (stim.start + 200)/h.dt

    amplitude_ctl = []
    dist_spine_ctl = []

    for i in range(1, N+1) :
        cellA = cell.cell(name="control")
        cellA.create_cell()

        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        cellA.set_cell_properties(sf=2.0)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amplitude_ctl.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))
        dist_spine_ctl.append(h.distance(cellA.soma(1), cellA.spine_necks[0](0)))

        del cellA


    amplitude_fcd = []
    dist_spine_fcd = []
    for i in range(1, N+1) :
        cellA = cell.cell(name="epileptogenic")
        cellA.create_cell()
        ampa_ctl = h.NetCon(stim, cellA.ampar[0])
        ampa_ctl.weight[0] = 31     #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cellA.nmdar[0])
        nmda_ctl.weight[0] = 14     #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        cellA.set_cell_properties(sf=1.2)

        h.finitialize(vinit)
        h.continuerun(tstop)      

        amplitude_fcd.append(cellA.v_soma.max(compute_start, compute_end) - cellA.v_soma.min(compute_start, compute_end))
        dist_spine_fcd.append(h.distance(cellA.soma(1), cellA.spine_necks[0](0)))

        del cellA


    min_amp = min(min(amplitude_ctl), min(amplitude_fcd))
    max_amp = max(max(amplitude_ctl), max(amplitude_fcd))
    # normalize amplitude
    amplitude_ctl_cvt = [(v - min_amp) / (max_amp - min_amp) * 300 for v in amplitude_ctl]
    amplitude_fcd_cvt = [(v - min_amp) / (max_amp - min_amp) * 300 for v in amplitude_fcd]

    fig, ax = plt.subplots(figsize=(3,4))
    fig.subplots_adjust(left=0.2)


    np.random.seed(1) ## 재생성을 위한 시드 넘버
 
    x_jitter1 = np.random.normal(1, 0.3, size=len(dist_spine_ctl))
    x_jitter2 = np.random.normal(1, 0.3, size=len(dist_spine_fcd))+1
    
    sc= plt.scatter(x_jitter1, dist_spine_ctl, marker="o", s=amplitude_ctl_cvt, color="cyan", alpha=0.3, label="Baseline")
    sc= plt.scatter(x_jitter2, dist_spine_fcd, marker="o", s=amplitude_fcd_cvt, color="magenta", alpha=0.3, label="Altered")
    
    kw = dict(prop="sizes", num=2, color="grey", fmt="{x:.2f}mV",
          func=lambda s: np.sqrt(s/500)*(max_amp - min_amp)+min_amp)
    legend2 = ax.legend(*sc.legend_elements(**kw),
                    loc="upper center", ncol=2, fontsize=8, frameon=False)
    ax.add_artist(legend2)

    plt.ylabel("Distance from Soma (µm)")  
    
    plt.xlim((0.2, 2.8))
    plt.ylim((15, 440))

    ax.set_yticks(ticks=[30, 100, 200, 300, 380], labels=["30", "100", "200", "300", "380"])
    ax.set_xticks(ticks=[1,2], labels=["Baseline", "Altered"])
    plt.tick_params(
        axis='y',          # changes apply to the y-axis
        which='both',      # both major and minor ticks are affected
        right=False 
    )
    plt.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        bottom=True,      # ticks along the bottom edge are off
        top=False)#),         # ticks along the top edge are off

    plt.savefig(f"distance_vs_soma_amp_scatter.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    if show :
        plt.show()


def Ri_bothModel_various_dendrite_comb_bar(min_basal=1, max_basal=10) :
    from neuron import h
    from neuron.units import ms, mV
    import cell_singlespine_randomloc as cellA
    import numpy as np

    from singleCell.utils import measure_input_resistance

    h.load_file("stdrun.hoc")
    vinit = -70 * mV
    h.celsius = 37
    h.dt = 0.1 *ms
    h.v_init = vinit

    stim = h.NetStim()
    stim.start = 300 *ms
    stim.number = 1

    x_ctl = []
    y_ctl = []
    for nbasal in range(min_basal, max_basal+1) :
        x_ctl.append(nbasal)

        cell_ = cellA.cell(name="control")
        spec_ = dict()
        spec_["nbasal"] = nbasal

        cell_.create_cell_with_spec(spec_)

        ampa_ctl = h.NetCon(stim, cell_.ampar[0])
        ampa_ctl.weight[0] = 31 #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cell_.nmdar[0])
        nmda_ctl.weight[0] = 14 #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        #mesure RI with varius factor 
        cell_.set_cell_properties(sf=2.0)
        ri = measure_input_resistance(cell_)
        y_ctl.append(ri)


        h.finitialize(vinit)
        h.continuerun(tstop)

        del cell_

    x_fcd = []
    y_fcd = []
    for nbasal in range(min_basal, max_basal+1) :
        x_fcd.append(nbasal)

        cell_ = cellA.cell(name="epilepsy")
        spec_ = dict()
        spec_["nbasal"] = nbasal

        cell_.create_cell_with_spec(spec_)

        ampa_ctl = h.NetCon(stim, cell_.ampar[0])
        ampa_ctl.weight[0] = 31 #ampa_weight
        ampa_ctl.delay = 0 * ms
        nmda_ctl = h.NetCon(stim, cell_.nmdar[0])
        nmda_ctl.weight[0] = 14 #nmda_weight
        nmda_ctl.delay = 0 * ms

        tstop = 1_000 * ms

        #mesure RI with varius factor 
        cell_.set_cell_properties(sf=1.2)
        ri = measure_input_resistance(cell_)
        y_fcd.append(ri)


        h.finitialize(vinit)
        h.continuerun(tstop)

        del cell_

    fig, ax = plt.subplots(figsize=(2,2))

    bplot = ax.boxplot([y_ctl, y_fcd], widths=0.5, medianprops=dict(color="black"), patch_artist=True)  # will be used to label x-ticks

    # fill with colors
    for patch, color in zip(bplot['boxes'], ["cyan", "magenta"]):
        patch.set_facecolor(color)
    plt.ylim((80, 170))
    ax.set_yticks(ticks=[100, 140], labels=["100", "140"], fontsize=12)
    plt.xticks([])
    plt.savefig(f"ri_box.png", dpi=600, bbox_inches='tight', pad_inches=0.01, transparent=False)

    plt.show()

distance_from_soma_vs_somatic_amplitude(N=100, show=True, legend=False)
