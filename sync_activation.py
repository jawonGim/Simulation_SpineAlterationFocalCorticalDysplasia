import time
import random
import numpy as np
import os

from neuron import gui, h
from neuron.units import ms, mV

from matplotlib import pyplot as plt
from openpyxl import Workbook, load_workbook

import morphology as morphology

start_time = time.perf_counter()

def stim_nspines(cell, nspines, ampa_weight, nmda_weight, timing=100):
    #select random spines to stimulate
    random.seed(time.time())

    selected_spines = set()
    while 1 :
        selected_spines.add(random.randrange(0, len(cell.spine_heads), 1))
        if len(selected_spines) >= nspines :
            break
    # print(f"selected spines: {selected_spines}")

    #set stimulation 
    stims = []
    ncl_ampa = []
    ncl_nmda = []
    for i in range(len(cell.spine_heads)) :
        if i in selected_spines :
            stims.append(h.NetStim())
            stims[-1].start = timing * ms
            stims[-1].number = 1 
            stims[-1].noise = 0

            ncl_ampa.append(h.NetCon(stims[-1], cell.ampar[i]))    
            ncl_ampa[-1].weight[0] = ampa_weight
            ncl_ampa[-1].delay = 0 * ms
    
            ncl_nmda.append(h.NetCon(stims[-1], cell.nmdar[i]))
            ncl_nmda[-1].weight[0] = nmda_weight
            ncl_nmda[-1].delay = 0 * ms
    
    
    #run simulation
    h.finitialize(vinit)
    h.continuerun(tstop)

    return cell

def sync_stim_and_saveas_xlsx(model_) :
    sheet_name = model_
    xls_fname = "spine_stim_nspikes_small.xlsx"
    if os.path.isfile(xls_fname):
        wb = load_workbook(filename = xls_fname)
    else:
        wb = Workbook()
    ws = wb.create_sheet(sheet_name)

    ridx = 0


    for nact_spine in range(40, 501, 20):
    # for nact_spine in range(1, 81, 1):
        print(f"nact_spine: {nact_spine}")
        ridx += 1
        ws.cell(row=ridx, column=1, value=nact_spine)
        spike_sum = 0
        array  = []
        for ntrial in range(1, 21) :
            print(f"trial: {ntrial}")
            cell = morphology.cell(verbose=False)

            if model_ == "ctl" :
                cell.create_cell()
                weight = [31, 14] #[650, 288]  # [ampa_weight, nmda_weight]
            else :
                cellspec = dict()
                cellspec["nspines_perdend"] = 30
                cellspec["spine_len"] = 1.696
                cellspec["spine_head_diam"] = 0.718
                cellspec["spine_neck_diam"] = 0.301 
                cellspec["model"] = "epilepsy" #control or epilepsy
                cell.create_cell_with_spec(cellspec)
                weight = [82, 36] #[965, 428]  # [ampa_weight, nmda_weight]

            cell.set_cell_properties()
            cell = stim_nspines(cell, nact_spine, weight[0], weight[1])    
        
            nspike = len(list(cell.spike_times))
            array.append(nspike)
            ws.cell(row=ridx, column=ntrial+1, value=nspike)
            print(f"{nact_spine}spine(trial{ntrial}), firing {nspike}times")

        # ws.cell(row=ridx, column=2, value=ntrial)  #write number of trials
        ws.cell(row=ridx, column=ntrial+2, value=np.sum(array))  #write total number of spikes
        ws.cell(row=ridx, column=ntrial+3, value=np.mean(array))  #write mean
        ws.cell(row=ridx, column=ntrial+4, value=np.std(array))  #write stdev 

        wb.save(xls_fname)  #save often to avoid sad accident 


def sync_stim_and_potential_plot(model_="ctl", nact_spine=10, ntrial=10):
    print(f"nact_spine: {nact_spine}")

    for i in range(0, ntrial) :
        cell = morphology.cell(verbose=False)

        if model_ == "ctl" :
            cell.create_cell()
            weight = [31,14]   #[650, 288]  # [ampa_weight, nmda_weight]
        else :
            cellspec = dict()
            cellspec["nspines_perdend"] = 30
            cellspec["spine_len"] = 1.696
            cellspec["spine_head_diam"] = 0.718
            cellspec["spine_neck_diam"] = 0.301 
            cellspec["model"] = "epilepsy" #control or epilepsy
            cell.create_cell_with_spec(cellspec)
            weight = [82, 36]   #[965, 428]  # [ampa_weight, nmda_weight]

        cell.set_cell_properties()

        t = h.Vector()
        t.record(h._ref_t)

        cell = stim_nspines(cell, nact_spine, weight[0], weight[1])    
        
        plt.plot(t, cell.v_soma,linewidth=1)
    plt.xlabel('t (ms)')
    plt.ylabel('v (mV)')
    plt.show()
        


model_ = "ctl"
if __name__ == "__main__" :
    if len(sys.argv) > 1 : 
        if sys.argv[1] == "fcd" :
            model_ = "fcd"
        
        if len(sys.argv) > 2 :
            nspines = int(sys.argv[2])

print(f"Model: {model_}")

###
vinit = -70 * mV
vleak = vinit
celsius  = 37

tstop = 500 * ms
h.celsius = celsius
h.dt = 0.1 *ms
#### 

# sync_stim_and_saveas_xlsx(model_)
sync_stim_and_potential_plot(model_, nact_spine=nspines)


end_time = time.perf_counter()
print(f"Execution time: {end_time - start_time:.2f} sec")